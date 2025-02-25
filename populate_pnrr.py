from openpyxl import load_workbook
from utils import *

def populate_pnrr(template_file_wb, output_file, mese_amm_prin, selected_month):
    try:

        template_file_sheets = template_file_wb.sheetnames
        template_file_sheet = template_file_wb[template_file_sheets[0]]

        mese_amm_prin_wb = load_workbook(mese_amm_prin)
        mese_amm_prin_sheets = mese_amm_prin_wb.sheetnames

        months = ["Gennaio", "Febbraio", "Marzo",
                  "Aprile", "Maggio", "Giugno",
                  "Luglio", "Agosto", "Settembre",
                  "Ottobre", "Novembre", "Dicembre"]

        months_in_number = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
        periodsStart = [
            "01/01", "01/02", "01/03", "01/04", "01/05", "01/06", "01/07", "01/08", "01/09", "01/10", "01/11", "01/12",
        ]
        periodsEnd = [
            "/01", "/02", "/03", "/04", "/05", "/06", "/07", "/08", "/09", "/10", "/11", "/12",
        ]
        #
        sheet_number = selected_month - 1
        # fixed data
        mese_amm_prin_sheet = mese_amm_prin_wb.worksheets[sheet_number]
        project_number = mese_amm_prin_sheet.cell(row=5, column=9).value
        month = months[sheet_number]
        year = mese_amm_prin_sheets[sheet_number].split("-", 1)[-1].strip()
        array_of_wp = extract_unique_wp(mese_amm_prin_sheet, 15, 3)
        employee_name = mese_amm_prin_sheet.cell(row=8, column=9).value
        employee_surname = mese_amm_prin_sheet.cell(row=8, column=31).value
        employee_signature = ""
        date_of_employee_signature = ""

        for i in range(1, 999):
            value = mese_amm_prin_sheet.cell(row=i, column=1).value
            if value is not None and value.startswith("Data:"):
                date_of_employee_signature = value.split("Data:")[1].strip()
                employee_signature = mese_amm_prin_sheet.cell(row=i + 1, column=1).value.split("Firma:")[1].strip()
                break

        for i in range(1, 999):
            value = template_file_sheet.cell(row=i, column=1).value
            if value == "Data e firma dell'addetto al progetto":
                template_file_sheet.cell(row=i + 1, column=1, value=date_of_employee_signature + ", " + employee_signature)
                break

        template_file_sheet.cell(row=3, column=2, value=project_number)
        template_file_sheet.cell(row=6, column=10, value=year)
        template_file_sheet.cell(row=23, column=2, value=month + " " + year)
        template_file_sheet.cell(row=10, column=3, value=employee_surname + " " + employee_name)
        template_file_sheet.title = months_in_number[sheet_number] + "-" + year

        insert_wp_rows(template_file_sheet, array_of_wp, row_copied=26)

        periodStart = periodsStart[sheet_number]
        tempPeriodEnd = ""

        for i in range(32):
            column = i + 17
            day = mese_amm_prin_sheet.cell(row=11, column=column).value
            if day == "Totale":
                break
            else:
                tempPeriodEnd = day
                is_ferial = is_ferial_day(day, month, year)
                if not is_ferial:
                    gray_fill = PatternFill(start_color="d8d8d8", end_color="d8d8d8", fill_type="solid")
                    col = i + 2
                    if len(array_of_wp) == 0:
                        len_for_cycle = 1
                    else:
                        len_for_cycle = len(array_of_wp)

                    # Colora le celle desiderate
                    for j in range(9+len_for_cycle):
                        row = 24+j
                        template_file_sheet.cell(row=row, column=col).fill = gray_fill

        periodEnd = tempPeriodEnd + periodsEnd[sheet_number]

        period_string_built = "Dal     " + periodStart + "/" + year + "                     al  " + periodEnd + "/" + year
        # # Insert the value into the template
        template_file_sheet.cell(row=2, column=2, value=period_string_built)

        update_total_hours_formula(template_file_sheet, array_of_wp, row_copied=26)
        update_total_hours_of_other_formula(template_file_sheet, array_of_wp, row_copied=26)
        populate_hours_from_mese(mese_amm_prin_sheet, template_file_sheet, start_row_wp=15, start_row_other=13, start_col_days=17, wp_output_start_row=26, day_row=11)

        # Save the updated workbook
        template_file_wb.save(output_file)
        print(f"Data populated and saved successfully to '{output_file}'.")

    except FileNotFoundError:
        print(f"Error: Template file '{template_file_wb}' or input file '{mese_amm_prin}' not found.")
    except Exception as e:
        print(f"Error while processing template file '{template_file_wb}': {e}")