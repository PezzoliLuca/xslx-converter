import openpyxl
from openpyxl import load_workbook
from utils import *

def populate_mic(template_file_wb, output_file, mese_amm_prin, start_month, end_month):
    try:
        months = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
        template_file_sheets = template_file_wb.sheetnames
        template_file_sheet = template_file_wb[template_file_sheets[0]]

        # Load the data from MESE_AMM_PRIN
        mese_amm_prin_wb = load_workbook(mese_amm_prin)

        # fixed data
        mese_amm_prin_sheet_0 = mese_amm_prin_wb.worksheets[0]
        year = mese_amm_prin_sheet_0.cell(row=1, column=48).value
        mese_amm_prin_sheet = mese_amm_prin_wb.worksheets[start_month - 1]
        project_title = mese_amm_prin_sheet.cell(row=3, column=9).value
        project_number = ""
        if mese_amm_prin_sheet.cell(row=5, column=9).value is not None:
            project_number = int(mese_amm_prin_sheet.cell(row=5, column=9).value)

        subject = mese_amm_prin_sheet.cell(row=6, column=9).value
        employee_name = mese_amm_prin_sheet.cell(row=8, column=9).value
        employee_surname = mese_amm_prin_sheet.cell(row=8, column=31).value
        employee_signature = ""
        date_of_employee_signature = ""
        supervisor_signature = ""
        date_of_supervisor_signature = ""


        for i in range(1, 999):
            value = mese_amm_prin_sheet.cell(row=i, column=1).value
            if value is not None and value.startswith("Data:"):
                date_of_employee_signature = value.split("Data:")[1].strip()
                employee_signature = mese_amm_prin_sheet.cell(row=i+1, column=1).value.split("Firma:")[1].strip()
                date_of_supervisor_signature = mese_amm_prin_sheet.cell(row=i, column=17).value.split("Data:")[1].strip()
                supervisor_signature = mese_amm_prin_sheet.cell(row=i + 1, column=17).value.split("Firma:")[1].strip()
                break

        for i in range(1, 999):
            value = template_file_sheet.cell(row=i, column=1).value
            if value == "Data e firma del personale":
                template_file_sheet.cell(row=i, column=1, value=value+ ": " + date_of_employee_signature + ", " + employee_signature)
                template_file_sheet.cell(row=i, column=17, value="Data e firma del supervisore"+ ": " + date_of_supervisor_signature + ", " + supervisor_signature)
                break

        # Insert the value into the template
        string_to_insert = str(project_title) + " - " + str(project_number)
        template_file_sheet.cell(row=1, column=13, value=string_to_insert)
        template_file_sheet.cell(row=2, column=13, value=str(subject))

        other_string_to_insert = ""
        if employee_name is not None:
            other_string_to_insert += employee_name
        if employee_surname is not None:
            other_string_to_insert += " " + employee_surname
        template_file_sheet.cell(row=3, column=13, value=other_string_to_insert)

        row_of_total = 0
        count_for_line = 0
        for sheet_number in range(start_month - 1, end_month):
            print("THE SHEET NUMBER IS: ", sheet_number)
            month = months[sheet_number]
            template_file_sheet.cell(row=5, column=2, value=year)
            template_file_sheet.cell(row=6+count_for_line*4, column=2, value=month)

            mese_amm_prin_sheet = mese_amm_prin_wb.worksheets[sheet_number]

            for i in range(1, 100):
                temp_value = mese_amm_prin_sheet.cell(row=i, column=1).value
                if temp_value == "Totale ore":
                    row_of_total = i
                    break

            row_of_output = 8 + count_for_line * 4
            # data with cycles
            for i in range(1, 32):
                is_ferial = is_ferial_day(str(i),str(month), str(year))
                if not is_ferial:
                    gray_fill = PatternFill(start_color="d8d8d8", end_color="d8d8d8", fill_type="solid")
                    col = i + 1
                    template_file_sheet.cell(row=7+count_for_line*4, column=col).fill = gray_fill
                    template_file_sheet.cell(row=8 + count_for_line * 4, column=col).fill = gray_fill


                column = i + 16
                if mese_amm_prin_sheet.cell(row=11, column=column).value == "Totale":
                    break

                hours = mese_amm_prin_sheet.cell(row=row_of_total, column=column).value
                if hours is None:
                    break
                hours_converted = convert_to_decimal_hours(hours)

                column_output = i + 1

                print("----->", row_of_output, column_output, hours_converted)
                template_file_sheet.cell(row=row_of_output, column=column_output, value=hours_converted)

            count_for_line += 1


        # Save the updated workbook
        template_file_wb.save(output_file)
        print(f"Data populated and saved successfully to '{output_file}'.")

    except FileNotFoundError:
        print(f"Error: Template file '{template_file_wb}' or input file '{mese_amm_prin}' not found.")
    except Exception as e:
        print(f"Error while processing template file '{template_file_wb}': {e}")