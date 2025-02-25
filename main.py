from populate_h2020 import *
from populate_mic import *
from populate_mur import *
from populate_pnrr import *
import PySimpleGUI as sg
from openpyxl import load_workbook


def select_file(prompt_message):
    sg.theme('DarkBlue3')  # Tema opzionale
    file_path = sg.popup_get_file(prompt_message, file_types=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))
    return file_path


def select_month_range(output_option):
    layout = [
        [sg.Text(f"Seleziona il mese per l'opzione {output_option}:")],
        [sg.Text("Mese (1-12):"), sg.InputText(size=(5, 1), key="month")],
        [sg.Button("Conferma"), sg.Button("Annulla")]
    ]

    window = sg.Window(f"Seleziona mese per {output_option}", layout, finalize=True)

    event, values = window.read()
    window.close()

    if event == sg.WINDOW_CLOSED or event == "Annulla":
        return None
    try:
        selected_month = int(values["month"])
        if 1 <= selected_month <= 12:
            return selected_month
        else:
            sg.popup_error("Selezione mese non valida! Inserisci un numero tra 1 e 12.")
            return None
    except ValueError:
        sg.popup_error("Input non valido. Inserisci un numero tra 1 e 12.")
        return None

def select_month_range_for_mic():
    layout = [
        [sg.Text("Seleziona il mese di inizio e di fine per l'opzione MIC:")],
        [sg.Text("Mese di inizio (1-12):"), sg.InputText(size=(5, 1), key="start_month")],
        [sg.Text("Mese di fine (1-12):"), sg.InputText(size=(5, 1), key="end_month")],
        [sg.Button("Conferma"), sg.Button("Annulla")]
    ]

    window = sg.Window("Seleziona mesi per MIC", layout, finalize=True)

    event, values = window.read()
    window.close()

    if event == sg.WINDOW_CLOSED or event == "Annulla":
        return None, None

    try:
        start_month = int(values["start_month"])
        end_month = int(values["end_month"])
        if 1 <= start_month <= 12 and 1 <= end_month <= 12 and start_month <= end_month:
            return start_month, end_month
        else:
            sg.popup_error(
                "Mesi non validi! Assicurati che i mesi siano tra 1 e 12 e che il mese di inizio sia prima di quello di fine.")
            return None, None
    except ValueError:
        sg.popup_error("Input non valido. Inserisci numeri tra 1 e 12.")
        return None, None


def select_output_option():
    layout = [
        [sg.Text("Seleziona l'opzione di output:")],
        [sg.Button("H2020"), sg.Button("MIC"), sg.Button("MUR"), sg.Button("PNRR")],
        [sg.Button("Esci")]
    ]

    window = sg.Window("Selezione tipo di output", layout, finalize=True)

    event, _ = window.read()
    window.close()

    if event == "H2020":
        return "H2020"
    elif event == "MIC":
        return "MIC"
    elif event == "MUR":
        return "MUR"
    elif event == "PNRR":
        return "PNRR"
    else:
        sg.popup("Operazione annullata. Uscita.")
        return None


def select_file_for_option(option):
    if option in ["H2020", "MUR"]:
        return select_file("Seleziona il file 'AMM_NOTE_PROG' Excel")
    elif option in ["MIC", "PNRR"]:
        return select_file("Seleziona il file 'MESE_AMM_PRIN' Excel")
    return None


def populate_template(template_file, output_file, amm_note_prog, mese_amm_prin, output_option):
    try:
        template_file_wb = load_workbook(template_file)
        if output_option == "H2020":
            populate_h2020(template_file_wb, output_file, amm_note_prog)

        elif output_option == "MIC":
            start_month, end_month = select_month_range_for_mic()
            if start_month is None or end_month is None:
                return
            num_months = end_month - start_month + 1  # Calcola il numero di mesi selezionati
            template_file = f"./templates/ESEMPIO_TS_MIC_{num_months}.xlsx"
            try:
                template_file_wb = load_workbook(template_file)
            except FileNotFoundError:
                sg.popup_error(
                    f"Attenzione: Il template '{template_file}' non esiste. Verrà utilizzato quello di default.")
                template_file_wb = load_workbook("./templates/ESEMPIO_TS_MIC_12.xlsx")

            populate_mic(template_file_wb, output_file, mese_amm_prin, start_month, end_month)

        elif output_option == "MUR":
            selected_month = select_month_range(output_option)
            if selected_month is None:
                return
            month_duration = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            n_days_in_month = month_duration[selected_month - 1]

            if n_days_in_month == 29:
                template_file = "./templates/ESEMPIO_TS_MUR_29gg.xlsx"
                template_file_wb = load_workbook(template_file)
            elif n_days_in_month == 30:
                template_file = "./templates/ESEMPIO_TS_MUR_30gg.xlsx"
                template_file_wb = load_workbook(template_file)

            populate_mur(template_file_wb, output_file, amm_note_prog, selected_month)

        elif output_option == "PNRR":
            selected_month = select_month_range(output_option)
            if selected_month is None:
                return
            month_duration = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            n_days_in_month = month_duration[selected_month - 1]

            if n_days_in_month == 29:
                template_file = "./templates/ESEMPIO_TS_PNRR_29gg.xlsx"
                template_file_wb = load_workbook(template_file)
            elif n_days_in_month == 30:
                template_file = "./templates/ESEMPIO_TS_PNRR_30gg.xlsx"
                template_file_wb = load_workbook(template_file)

            populate_pnrr(template_file_wb, output_file, mese_amm_prin, selected_month)

        else:
            sg.popup_error("Opzione non riconosciuta.")

    except FileNotFoundError:
        sg.popup_error(f"Errore: Il file del template '{template_file}' o del file di input non è stato trovato.")
    except Exception as e:
        sg.popup_error(f"Errore durante l'elaborazione del template '{template_file}': {e}")


def main():
    output_option = select_output_option()
    if not output_option:
        return

    template_files = {
        "H2020": "./templates/ESEMPIO_TS_H2020.xlsx",
        "MIC": "./templates/ESEMPIO_TS_MIC_12.xlsx",
        "MUR": "./templates/ESEMPIO_TS_MUR_31gg.xlsx",
        "PNRR": "./templates/ESEMPIO_TS_PNRR_31gg.xlsx"
    }

    if output_option in template_files:
        template_file = template_files[output_option]
        output_file = f"output_{output_option.lower()}.xlsx"

        mese_amm_prin = amm_note_prog = None

        if output_option in ["H2020", "MUR"]:
            amm_note_prog = select_file_for_option(output_option)

        if output_option in ["MIC", "PNRR"]:
            mese_amm_prin = select_file_for_option(output_option)

        populate_template(template_file, output_file, amm_note_prog, mese_amm_prin, output_option)


if __name__ == "__main__":
    main()
