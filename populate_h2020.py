import openpyxl
from openpyxl import load_workbook
from utils import *
from openpyxl.utils import get_column_letter
from copy import copy
import string

def populate_h2020(template_file_wb, output_file, amm_note_prog):
    try:
        template_file_sheets = template_file_wb.sheetnames
        template_file_sheet_data = template_file_wb[template_file_sheets[12]]

        # Load the data from AMM_NOTE_PROG
        amm_note_prog_wb = load_workbook(amm_note_prog)
        amm_note_prog_sheets = amm_note_prog_wb.sheetnames
        number_of_sheets = len(amm_note_prog_sheets)

        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        for sheet_number in range(number_of_sheets):
            print("THE SHEET NUMBER IS: ", sheet_number)
            # fixed data
            amm_note_prog_sheet = amm_note_prog_wb.worksheets[sheet_number]
            project_title = amm_note_prog_sheet.cell(row=4, column=2).value
            project_number = amm_note_prog_sheet.cell(row=5, column=2).value.split('—')[0].strip()
            employee = amm_note_prog_sheet.cell(row=7, column=2).value
            employee_signature = ""
            date_of_employee_signature = ""
            supervisor_signature = ""
            date_of_supervisor_signature = ""

            for i in range(1, 999):
                value = amm_note_prog_sheet.cell(row=i, column=1).value
                if value == "Employee Signature:":
                    employee_signature = amm_note_prog_sheet.cell(row=i, column=2).value
                    date_of_employee_signature = amm_note_prog_sheet.cell(row=i+1, column=2).value
                    supervisor_signature = amm_note_prog_sheet.cell(row=i, column=8).value
                    date_of_supervisor_signature = amm_note_prog_sheet.cell(row=i+1, column=8).value
                    break

            # Insert the value into the template
            template_file_sheet_data.cell(row=2, column=17, value=project_title)
            template_file_sheet_data.cell(row=2, column=18, value=project_number)

            template_file_sheet = template_file_wb[template_file_sheets[sheet_number]]
            template_file_sheet.cell(row=7, column=9, value=employee)

            for i in range(1, 999):
                value = template_file_sheet.cell(row=i, column=1).value
                if value == "Signature (Name of the person working for the action)":
                    template_file_sheet.cell(row=i, column=9, value=employee_signature)
                    template_file_sheet.cell(row=i+1, column=9, value=date_of_employee_signature)
                    template_file_sheet.cell(row=i, column=29, value=supervisor_signature)
                    template_file_sheet.cell(row=i+1, column=29, value=date_of_supervisor_signature)
                    break

            month = months[sheet_number]

            year = amm_note_prog_sheets[sheet_number].split("-", 1)[-1].strip()

            array_of_wp = extract_unique_wp(amm_note_prog_sheet)
            insert_wp_rows(template_file_sheet, array_of_wp, row_copied=10)
            update_total_hours_formula(template_file_sheet, array_of_wp, row_copied=10)
            populate_hours_from_amm(amm_note_prog_sheet, template_file_sheet, month, year, start_row_input=11, start_row_output=10)

            template_file_sheet.cell(row=3, column=24, value=year)
            template_file_sheet.cell(row=3, column=31, value=month)

        # Save the updated workbook
        template_file_wb.save(output_file)
        print(f"Data populated and saved successfully to '{output_file}'.")

    except FileNotFoundError:
        print(f"Error: Template file '{template_file_wb}' or input file '{amm_note_prog}' not found.")
    except Exception as e:
        print(f"Error while processing template file '{template_file_wb}': {e}")