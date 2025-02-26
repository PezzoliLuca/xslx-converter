from openpyxl.styles import PatternFill
from datetime import datetime
from openpyxl.utils import get_column_letter
from copy import copy

def is_ferial_day(day: str, month: str, year: str) -> bool:
    month_map = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,

        "Gennaio": 1, "Febbraio": 2, "Marzo": 3, "Aprile": 4, "Maggio": 5, "Giugno": 6,
        "Luglio": 7, "Agosto": 8, "Settembre": 9, "Ottobre": 10, "Novembre": 11, "Dicembre": 12
    }

    italian_festivity = {
        "01-01", "06-01", "25-04", "01-05", "02-06", "15-08",
        "01-11", "08-12", "25-12", "26-12"
    }

    try:
        month_number = month_map.get(month)

        if month_number is None:
            return True

        date = datetime(int(year), month_number, int(day))
        date_str = date.strftime("%d-%m")

        return date.weekday() < 5 and date_str not in italian_festivity
    except ValueError as e:
        print(f"Error in creation of the date: {e}")  # Debug
        return True

def convert_to_decimal_hours(time_str):
    hours, minutes = map(int, time_str.split(':'))
    decimal_hours = hours + minutes / 60
    return decimal_hours

def number_to_letters(n):
    result = ""
    while n > 0:
        n -= 1  # Shift per rendere A=1, B=2...
        result = chr(65 + (n % 26)) + result
        n //= 26
    return result

def extract_unique_wp(sheet, start_row=11, project_col=8):
    """
    Extract uniques WP form xslx sheet starting from the `start_row` and taking the value from the `project_col`.
    """
    unique_wp = set()
    array_of_wp = []

    for i in range(999):
        row = i + start_row
        if project_col == 8 and sheet.cell(row=row, column=1).value is None:
            break
        if project_col == 3 and sheet.cell(row=row, column=project_col).value is None:
            break

        working_project = sheet.cell(row=row, column=project_col).value
        if working_project not in unique_wp and working_project is not None:
            unique_wp.add(working_project)
            array_of_wp.append(working_project)

    return array_of_wp


def insert_wp_rows(template_sheet, array_of_wp, row_copied, col_start=1):
    """
    Insert the rows for each WP and copies the data from the `row_copied`.
    """
    if not array_of_wp:
        return

    col_end = template_sheet.max_column
    merged_cells = list(template_sheet.merged_cells)

    # Rimuove unioni delle celle sotto la riga copiata
    for merged_cell in merged_cells:
        if merged_cell.min_row > row_copied:
            template_sheet.unmerge_cells(str(merged_cell))

    try:
        # Sposta in basso tutte le righe sotto row_copied
        template_sheet.move_range(
            f"A{row_copied + 1}:{get_column_letter(col_end)}{template_sheet.max_row}",
            rows=len(array_of_wp) - 1,
            cols=0
        )
    except Exception as e:
        print(f"Errore nello spostamento delle righe: {e}")

    # Ripristina le unioni delle celle dopo lo spostamento
    for merged_cell in merged_cells:
        if merged_cell.min_row >= row_copied - 1:
            new_min_row = merged_cell.min_row + len(array_of_wp) - 1
            new_max_row = merged_cell.max_row + len(array_of_wp) - 1
            template_sheet.merge_cells(
                start_row=new_min_row,
                start_column=merged_cell.min_col,
                end_row=new_max_row,
                end_column=merged_cell.max_col
            )

    # Inserisce le righe dei WP
    for i, wp in enumerate(array_of_wp):
        new_row = row_copied + i
        for col in range(col_start, col_end + 1):
            cell_orig = template_sheet.cell(row=row_copied, column=col)
            cell_new = template_sheet.cell(row=new_row, column=col)

            if cell_orig.value is not None:
                if cell_orig.data_type == "f":
                    formula = str(cell_orig.value).replace(str(row_copied), str(new_row))
                    cell_new.value = formula
                else:
                    cell_new.value = cell_orig.value

            if cell_orig.has_style:
                cell_new._style = copy(cell_orig._style)

        # Inserisce il valore del WP
        template_sheet.cell(row=new_row, column=1, value=wp)


def update_total_hours_formula(template_sheet, array_of_wp, row_copied, total_hours_row_offset=1, col_start=1):
    """
    Aggiorna la formula della riga "Total hours" dopo l'inserimento dei WP.
    """
    col_end = template_sheet.max_column
    total_row = row_copied + len(array_of_wp)

    for col in range(col_start, col_end + 1):
        cell = template_sheet.cell(row=total_row, column=col)
        if cell.value is not None and cell.data_type == "f":
            letter = get_column_letter(col)
            if total_row > row_copied:
                formula = str(cell.value).replace(
                    str(row_copied), f"{row_copied}:{letter}{total_row - 1}"
                )
                cell.value = formula

def update_total_hours_of_other_formula(template_sheet, array_of_wp, row_copied, col_start=1):
    """
    Update the formula of "Total hours" after the insert of WPs.
    """
    col_end = template_sheet.max_column
    total_row = row_copied + len(array_of_wp)

    for row in range(total_row + 1, total_row + 1 + 6):
        for col in range(col_start, col_end + 1):
            cell = template_sheet.cell(row=row, column=col)
            if cell.value is not None and cell.data_type == "f":
                if row < row_copied + len(array_of_wp) + 6:
                    formula = str(cell.value).replace(
                        str(row - len(array_of_wp) + 1), f"{row}"
                    )
                    cell.value = formula
                else:
                    formula = str(cell.value).replace(
                        str(row_copied + 1), f"{total_row}:{get_column_letter(col)}{total_row+5}"
                    )
                    cell.value = formula


def populate_hours_from_amm(sheet_input, sheet_output, month, year, start_row_input=11, start_row_output=9, day_col=1, project_col=8, hours_col=10):
    """
    Popola le ore dal foglio `sheet_input` al foglio `sheet_output`.
    """
    for i in range(999):
        row = i + start_row_input
        day = sheet_input.cell(row=row, column=day_col).value
        if day is None:
            break

        working_project = sheet_input.cell(row=row, column=project_col).value
        hours = sheet_input.cell(row=row, column=hours_col).value
        hours_converted = convert_to_decimal_hours(hours)

        is_ferial = is_ferial_day(day, month, year)
        if working_project is None and is_ferial:
            continue

            # Se la data non è feriale, colora tutta la colonna relativa al giorno
        if not is_ferial:
            gray_fill = PatternFill(start_color="d8d8d8", end_color="d8d8d8", fill_type="solid")
            # Scorri tutte le righe fino a "Total hours "
            for j in range(999):
                row_output = j + start_row_output
                wp_name = sheet_output.cell(row=row_output, column=1).value
                # Colora la cella desiderata
                sheet_output.cell(row=row_output - 1, column=int(day) + 2).fill = gray_fill
                sheet_output.cell(row=row_output, column=int(day) + 2).fill = gray_fill
                # Se trovi la riga "Total hours ", interrompi il ciclo
                if wp_name == "Total hours ":
                    break

        for j in range(999):
            row_output = j + start_row_output
            wp_name = sheet_output.cell(row=row_output, column=1).value

            if wp_name == "Total hours ":
                break
            if wp_name == working_project:
                current_value = sheet_output.cell(row=row_output, column=int(day) + 2).value
                new_value = (current_value or 0) + hours_converted
                print("----->", row_output, int(day) + 2, new_value)
                sheet_output.cell(row=row_output, column=int(day) + 2, value=new_value)
                break

def populate_hours_from_mese(mese_amm_prin_sheet, template_file_sheet, start_row_wp=15, start_row_other=13, start_col_days=17, wp_output_start_row=26, day_row=11):
    """
    Popola le ore dal foglio 'mese_amm_prin_sheet' al foglio 'template_file_sheet'.

    :param mese_amm_prin_sheet: Foglio di input (mese amministrativo principale)
    :param template_file_sheet: Foglio di output (template)
    :param start_row_wp: Riga iniziale dei WP nel foglio di input
    :param start_row_other: Riga iniziale delle attivita' extra nel foglio di input
    :param start_col_days: Colonna iniziale dei giorni nel foglio di input
    :param wp_output_start_row: Riga iniziale dei WP nel foglio di output
    :param day_row: Riga dei giorni nel foglio di input
    """
    # data with cycles per WP
    for i in range(31):
        row = i + start_row_wp
        working_project = mese_amm_prin_sheet.cell(row=row, column=3).value

        if working_project is None:
            break

        for j in range(999):
            column = j + start_col_days
            value = mese_amm_prin_sheet.cell(row=row, column=column).value
            hours = "00:00"
            if value is not None:
                hours = value
            hours_converted = convert_to_decimal_hours(hours)

            row_output = 0
            for x in range(999):
                internal_row = x + wp_output_start_row
                wp_name = template_file_sheet.cell(row=internal_row, column=1).value
                if wp_name == "Tot ore progetto ":
                    break
                if wp_name == working_project:
                    row_output = internal_row
                    break

            if row_output != 0:
                day = mese_amm_prin_sheet.cell(row=day_row, column=column).value
                if day == "Totale":
                    break

                value_to_insert = hours_converted if hours_converted != 0.0 else ""
                print("----->", row_output, int(day) + 1, value_to_insert)
                template_file_sheet.cell(row=row_output, column=int(day) + 1, value=value_to_insert)

    # other data with cycles
    activity_mapping = {
        "Attività svolta su altri progetti": "Altri progetti finanziati",
        "Attività ordinaria": "Attività ordinaria",
        "Altro (Malattia, Ferie...)": "Malattia"
    }

    for i in range(37):
        row = i + start_row_other
        value = mese_amm_prin_sheet.cell(row=row, column=1).value

        if value in activity_mapping:
            for j in range(999):
                column = j + start_col_days
                hours_value = mese_amm_prin_sheet.cell(row=row, column=column).value
                hours = "00:00"
                if hours_value is not None:
                    hours = hours_value
                hours_converted = convert_to_decimal_hours(hours)

                row_output = 0
                for x in range(999):
                    internal_row = x + wp_output_start_row
                    activity = template_file_sheet.cell(row=internal_row, column=1).value
                    if activity == activity_mapping[value]:
                        row_output = internal_row
                        break

                if row_output != 0:
                    day = mese_amm_prin_sheet.cell(row=day_row, column=column).value
                    if day == "Totale":
                        break

                    value_to_insert = hours_converted if hours_converted != 0.0 else ""
                    template_file_sheet.cell(row=row_output, column=int(day) + 1, value=value_to_insert)
