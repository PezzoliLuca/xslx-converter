from openpyxl import load_workbook
from utils import *

def populate_mur(template_file_wb, output_file, amm_note_prog, selected_month):
    try:

        template_file_sheets = template_file_wb.sheetnames
        template_file_sheet = template_file_wb[template_file_sheets[0]]

        # Load the data from AMM_NOTE_PROG
        amm_note_prog_wb = load_workbook(amm_note_prog)
        amm_note_prog_sheets = amm_note_prog_wb.sheetnames

        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        sheet_number = selected_month - 1
        # fixed data
        amm_note_prog_sheet = amm_note_prog_wb.worksheets[sheet_number]
        project_title = amm_note_prog_sheet.cell(row=4, column=2).value
        project_cup = amm_note_prog_sheet.cell(row=4, column=13).value
        employee = amm_note_prog_sheet.cell(row=7, column=2).value
        employee_signature = ""
        date_of_employee_signature = ""
        supervisor_signature = ""
        date_of_supervisor_signature = ""

        for i in range(1, 999):
            value = amm_note_prog_sheet.cell(row=i, column=1).value
            if value == "Employee Signature:":
                employee_signature = amm_note_prog_sheet.cell(row=i, column=2).value
                date_of_employee_signature = amm_note_prog_sheet.cell(row=i + 1, column=2).value
                supervisor_signature = amm_note_prog_sheet.cell(row=i, column=8).value
                date_of_supervisor_signature = amm_note_prog_sheet.cell(row=i + 1, column=8).value
                break

        # Insert the value into the template
        template_file_sheet.cell(row=5, column=9, value=project_title)
        template_file_sheet.cell(row=5, column=28, value=project_cup)
        template_file_sheet.cell(row=6, column=9, value=employee)

        for i in range(1, 999):
            value = template_file_sheet.cell(row=i, column=1).value
            if value is not None:
                value = value.replace("\n", " ")
            if value == "Firma (Persona che ha lavorato nell'azione) Date: ":
                template_file_sheet.cell(row=i, column=9, value=employee_signature)
                template_file_sheet.cell(row=i + 1, column=9, value=date_of_employee_signature)
                template_file_sheet.cell(row=i, column=29, value=supervisor_signature)
                template_file_sheet.cell(row=i + 1, column=29, value=date_of_supervisor_signature)
                break

        month = months[sheet_number]
        year = amm_note_prog_sheets[sheet_number].split("-", 1)[-1].strip()

        array_of_wp = extract_unique_wp(amm_note_prog_sheet)
        insert_wp_rows(template_file_sheet, array_of_wp, row_copied=9)
        update_total_hours_formula(template_file_sheet, array_of_wp, row_copied=9)
        populate_hours_from_amm(amm_note_prog_sheet, template_file_sheet, month, year, start_row_input=11, start_row_output=9)

        template_file_sheet.cell(row=3, column=24, value=year)
        template_file_sheet.cell(row=3, column=34, value=month)

        # Save the updated workbook
        template_file_wb.save(output_file)
        print(f"Data populated and saved successfully to '{output_file}'.")

    except FileNotFoundError:
        print(f"Error: Template file '{template_file_wb}' or input file '{amm_note_prog}' not found.")
    except Exception as e:
        print(f"Error while processing template file '{template_file_wb}': {e}")